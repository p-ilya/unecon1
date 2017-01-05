import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string as col_ifs
from openpyxl.utils.exceptions import InsufficientCoordinatesException

filename = 'fekonomiki_i_finansov_ofo_1_kurs_0'
wb1 = load_workbook('{}.xlsx'.format(filename), read_only=False, guess_types=True)


def find_parse_start(ws):
    #  finding where the actual data is stored
    find_range = ws['A1:C20']
    for row in find_range:
        for cell in row:
            if str(cell.value).lower() == 'понедельник':
                col1 = cell.column
                row1 = cell.row
                break
            else:
                continue
        try:
            if row1: break
        except NameError:
            continue

    return ('{0}{1}'.format(col1, row1))



def unmerge_worksheet(ws):
    corner = find_parse_start(ws)

    for diapason in ws.merged_cell_ranges:
        first, last = diapason.split(':')
        data = ws[first].value
        try:
            ws.unmerge_cells(diapason)
            print('Unmerged: {0}'.format(diapason))

            #  copying the data
            col_num = col_ifs(ws[last].column) - col_ifs(ws[first].column)
            if col_num > 1 and 'A' not in first:
                place = ws.cell(row=ws[first].row, column=col_ifs(ws[first].column) - 1).value
                for i in range(0,col_num+1,2):
                    for cell_range in ws[first:last]:
                        cell_range[i].value = data
                for i in range(1,col_num+1,2):
                    for cell_range in ws[first:last]:
                        cell_range[i].value = place

            else:
                for cell_range in ws[first:last]:
                    for cell in cell_range:
                        cell.value = data

        except InsufficientCoordinatesException:
            print('These cells are not merged: {}'.format(diapason))
        '''
        try:
            print(ws[first:last])
        except ValueError as e:
            print('Error occured. Details:\n{}'.format(e))
        '''



for sheet in wb1:
    unmerge_worksheet(sheet)
    wb1.save('{}_NEW.xlsx'.format(filename))

